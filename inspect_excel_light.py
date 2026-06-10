import openpyxl
import os

def inspect_xlsx(filepath):
    print(f"--- Inspecting {os.path.basename(filepath)} ---")
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        print("Sheet names:", wb.sheetnames)
        
        # Print first 10 rows
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            print(row)
            if i >= 10:
                break
        print("\n")
    except Exception as e:
        print(f"Error reading {filepath}: {e}")

base_dir = '/Users/pmpmpm/Antigravity/passenger_capacity/extracted'
files_to_inspect = ['113年7月.xlsx']

for f in files_to_inspect:
    inspect_xlsx(os.path.join(base_dir, f))
